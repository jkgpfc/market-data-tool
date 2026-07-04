"""Excel export helpers for backtest outputs."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from backtest_engine import BacktestResult


def backtest_to_excel_bytes(result: BacktestResult) -> bytes:
    buffer = BytesIO()
    write_backtest_excel(result, buffer)
    return buffer.getvalue()


def write_backtest_excel(result: BacktestResult, output: str | Path | BytesIO) -> None:
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame([result.summary]).to_excel(writer, index=False, sheet_name="Summary")
        result.trade_log.to_excel(writer, index=False, sheet_name="Trade Log")
        result.mtm.to_excel(writer, index=False, sheet_name="MTM")
        result.equity_curve.to_excel(writer, index=False, sheet_name="Equity Curve")
        result.yearly_pnl.to_excel(writer, index=False, sheet_name="Yearly PnL")
        result.monthly_pnl.to_excel(writer, index=False, sheet_name="Monthly PnL")
        for worksheet in writer.book.worksheets:
            _style_sheet(worksheet)


def _style_sheet(worksheet) -> None:  # noqa: ANN001 - openpyxl worksheet type
    if worksheet.max_row == 0:
        return
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx in range(1, worksheet.max_column + 1):
        values = [worksheet.cell(row=row, column=col_idx).value for row in range(1, min(worksheet.max_row, 200) + 1)]
        width = min(max(len(str(value)) if value is not None else 0 for value in values) + 2, 45)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = width

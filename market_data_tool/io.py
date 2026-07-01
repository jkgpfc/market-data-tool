"""Input and output helpers."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .config import SYMBOL_COLUMN_CANDIDATES
from .models import MarketDataResult
from .symbols import SymbolRequest, normalize_symbol


def read_symbols_from_csv(path: str | Path) -> list[SymbolRequest]:
    """Read and normalize symbols from a CSV file.

    The preferred symbol column is `symbol`, with `ticker` and `instrument`
    accepted as aliases. If none are present, the first column is used. An
    optional `exchange` column can contain `NSE` or `BSE` to append the correct
    Yahoo Finance suffix.
    """
    csv_path = Path(path)
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    frame = pd.read_csv(csv_path)
    if frame.empty:
        return []

    lower_to_actual = {str(column).lower(): column for column in frame.columns}
    selected = next(
        (lower_to_actual[name] for name in SYMBOL_COLUMN_CANDIDATES if name in lower_to_actual),
        frame.columns[0],
    )
    exchange_column = lower_to_actual.get("exchange")
    requests: list[SymbolRequest] = []
    for index, value in frame[selected].dropna().items():
        symbol = str(value).strip()
        if not symbol:
            continue
        exchange = str(frame.at[index, exchange_column]).strip() if exchange_column and pd.notna(frame.at[index, exchange_column]) else None
        requests.append(normalize_symbol(symbol, exchange))
    return requests


def merge_symbols(
    cli_symbols: Iterable[str] | None,
    csv_symbols: Iterable[SymbolRequest] | None,
) -> list[SymbolRequest]:
    """Merge symbol sources while preserving order and de-duplicating."""
    merged: list[SymbolRequest] = []
    seen: set[str] = set()

    for raw_symbol in cli_symbols or []:
        request = normalize_symbol(str(raw_symbol))
        key = request.yahoo_symbol.upper()
        if request.yahoo_symbol and key not in seen:
            merged.append(request)
            seen.add(key)

    for request in csv_symbols or []:
        key = request.yahoo_symbol.upper()
        if request.yahoo_symbol and key not in seen:
            merged.append(request)
            seen.add(key)

    return merged


def export_to_excel(results: list[MarketDataResult], output_path: str | Path) -> Path:
    """Export results to a formatted Excel workbook."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    rows = [result.to_dict() for result in results]
    frame = pd.DataFrame(rows)
    column_order = [
        "original_symbol",
        "symbol",
        "status",
        "cmp",
        "vwap",
        "hourly_rsi_14",
        "daily_rsi_14",
        "weekly_rsi_14",
        "monthly_rsi_14",
        "currency",
        "exchange",
        "instrument_type",
        "error",
        "fetched_at_utc",
    ]
    frame = frame.reindex(columns=column_order)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, sheet_name="Market Data")
        worksheet = writer.sheets["Market Data"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 45)

    return path
